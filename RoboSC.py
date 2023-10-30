from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import time
import pandas as pd
import openpyxl
import re
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from datetime import datetime
import os
from selenium.webdriver.support.ui import Select

book = openpyxl.load_workbook(
    r"C:\Users\Gabriel V Costa\Desktop\chrome-win\BotSC.xlsx"
)
sheet = book.active
def clear_console():
    os.system('cls' if os.name == 'nt' else 'clear')
options = webdriver.ChromeOptions()
# Caminho para o ChromeDriver (caso esteja usando o ChromeDriverManager, não precisa desse caminho)
chrome_driver_path = r"C:\Users\Gabriel V Costa\Desktop\chrome-win\chromedriver.exe"

# Configurar as opções do Chrome
chrome_options = webdriver.ChromeOptions()

# Diretório para armazenar o cache
cache_dir = os.path.join(os.getcwd(), "chrome_cache")

# Criar o diretório se ele não existir
if not os.path.exists(cache_dir):
    os.makedirs(cache_dir)

# Configurar o perfil do ChromeDriver com o cache ativado
chrome_options.add_argument(f"--user-data-dir={cache_dir}")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--no-sandbox")

# Instanciar o serviço do ChromeDriver
s = Service(chrome_driver_path)

# Inicializar o navegador Chrome com o ChromeDriver e as opções configuradas
driver = webdriver.Chrome(service=s, options=chrome_options)

clear_console()
# Constantes
caminhoLocal = os.getcwd()  # diretório atual
driver.maximize_window()
totalSalvo  = 0
linkExterno = "https://eproc.jfsc.jus.br/eprocV2/externo_controlador.php"
# arquivoLocal = caminhoLocal + r"\index.html"  # página de exemplo
linkArquivo = r"C:\Users\Gabriel V Costa\Desktop\chrome-win"  # diretório do arquivo
def find_next_empty_row(sheet, column):
    row = 1
    while sheet.cell(row=row, column=column).value is not None:
        row += 1
    return row

def value_exists(sheet, column, value):
    row = 1
    while sheet.cell(row=row, column=column).value is not None:
        if sheet.cell(row=row, column=column).value == value:
            return True
        row += 1
    return False

driver.get(linkExterno)  # executar o navegador
time.sleep(5)
driver.find_element(By.XPATH, '//*[@id="main-menu"]/li[4]/a').click()
time.sleep(1)
driver.find_element(By.XPATH, '//*[@id="menu-ul-3"]/li[1]/a').click()
time.sleep(1)
#
tabela = pd.read_excel(r"C:\Users\Gabriel V Costa\Desktop\chrome-win\exSC.xlsx")


# DataEvento recente(tabela)
def get_event_value(element):
    try:
        return abs(500 - int(element.get_attribute("id").replace("trEvento", "")))
    except (AttributeError, ValueError):
        return float("inf")



for i, cpf in enumerate(tabela["CPF"]):
    select = Select(driver.find_element(By.ID, "selTipoPesquisa"))
    select.select_by_visible_text("CPF/CNPJ")
    input_cpf = driver.find_element(By.XPATH, '//*[@id="divStrDocParte"]/dl/dd/input')
    input_cpf.send_keys(cpf)
    time.sleep(3)

    try:
        # Espera até 10 segundos até que o botão esteja disponível
        button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'sbmConsultar')))
        driver.execute_script("arguments[0].click();", button)
    except TimeoutException:
        # Se o botão não estiver disponível após 10 segundos, tente outra maneira
        print("Botão sbmConsultar não encontrado, tentando de outra maneira...")
        try:
            # Tente localizar o botão com um seletor NAME 
            button = driver.find_element(By.NAME, 'sbmConsultar')
            driver.execute_script("arguments[0].click();", button)
        except Exception as e:
            print(f"Erro ao tentar de outra maneira: {e}")

    time.sleep(2)
    try:
        alert = driver.switch_to.alert
        alert.accept()
        driver.switch_to.default_content()  
    except NoAlertPresentException:
        pass            
    except Exception as e:
        print("Ocorreu um erro:", str(e))
            

    try:
        
        # Checa se foi redirecionado a outra página
        # Se redirecionado, vá para o próximo CPF da lista
        driver.find_element(By.ID, "txtNumProcesso")
    except NoSuchElementException:
        # Limpa o campo e coloca o próximo CPF
        input_cpf.clear()
        continue
    # Dados
    
    numeroProcesso = driver.find_element(By.ID, "txtNumProcesso").text
    competencia = driver.find_element(By.ID, "txtCompetencia").text
    dtAutuacao = driver.find_element(By.ID, "txtAutuacao").text

    try:
        subsecao = driver.find_element(By.ID, "txtLocalidade").text
    except NoSuchElementException:
        subsecao = "N/A"
        print("Localidade não encontrada")
    juiz = driver.find_element(By.ID, "txtMagistrado").text

    driver.find_element(By.XPATH, '//*[@id="fldAssuntos"]/legend').click()
    descricaoAssunto = driver.find_element(
        By.XPATH, '//*[@id="conteudoAssuntos2"]/table/tbody/tr[2]/td[2]'
    ).text

    autor = driver.find_element(By.CLASS_NAME, "infraNomeParte").text

    advogados_element = driver.find_element(By.CLASS_NAME, "autorReu")
    advogados_text = advogados_element.text

    advogados_lines = advogados_text.split("\n")[1:]
    advogados_text = ", ".join([line.strip() for line in advogados_lines])
    advogados_text = re.sub(r"\b[A-Z]{2}\d+\b", "", advogados_text)
    advogados = advogados_text.strip(", ")

    reu = driver.find_element(By.ID, "spnNomeParteReu0").text
    time.sleep(0.5)
    event_elements = driver.find_elements(
        By.XPATH, '//*[starts-with(@id, "trEvento")]/td[2]'
    )
    # Find the element with the closest value to trEvento200
    closest_element = min(event_elements, key=get_event_value)
    # Get the data from the closest element
    ultimoMovimento = closest_element.text
    ultimoMovimentoDescricao = driver.find_element(By.CLASS_NAME, "infraEventoDescricao").text
    time.sleep(2)
    driver.find_element(By.XPATH, '//*[@id="btnNova"]/span').click()

    time.sleep(1)
    # Find the next available row to fill in the Excel sheet
    
    row = find_next_empty_row(sheet, 1)

    if not value_exists(sheet, 1, numeroProcesso):
        sheet.cell(row=row, column=1, value=numeroProcesso)
        sheet.cell(row=row, column=2, value=competencia)
        sheet.cell(row=row, column=3, value=dtAutuacao)
        sheet.cell(row=row, column=4, value=subsecao)
        sheet.cell(row=row, column=5, value=juiz)
        sheet.cell(row=row, column=6, value=descricaoAssunto)
        sheet.cell(row=row, column=7, value=reu)
        sheet.cell(row=row, column=8, value=ultimoMovimento)
        sheet.cell(row=row, column=9, value=ultimoMovimentoDescricao)
        sheet.cell(row=row, column=10, value=advogados)
        sheet.cell(row=row, column=11, value="SC")
        current_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        sheet.cell(row=row, column=12, value=current_time)
        sheet.cell(row=row, column=13, value=autor)
        sheet.cell(row=row, column=14, value="COM RESULTADO")
    
    #Salve no Excel após cada interação
    book.save(r"C:\Users\Gabriel V Costa\Desktop\chrome-win\BotSC.xlsx")
    totalSalvo +=1
    print (f"CPF: {cpf} cadastrado! Total: {totalSalvo}")
    time.sleep(0)

driver.quit()
